import streamlit as st
import json
import os
from datetime import datetime
from openpyxl import load_workbook

# ------------------ CONFIGURACIÓN ------------------ #
st.set_page_config(page_title="GeoProtocolos", layout="wide", initial_sidebar_state="expanded")

# ------------------ ESTILO PERSONALIZADO ------------------ #
st.markdown("""
    <style>
        html, body, [class*="css"] {
            font-family: 'Segoe UI', sans-serif;
            color: #E1E1E1;
        }
        .sidebar .sidebar-content {
            background-color: #1e1e2f;
        }
        .css-1d391kg { background-color: #1e1e2f !important; }
        .css-1v3fvcr { color: #E1E1E1 !important; font-weight: 600; }
        .block-container {
            padding: 2rem 3rem 3rem;
            background-color: #12121b;
        }
        h1, h2, h3 {
            color: #F0F0F0;
        }
    </style>
""", unsafe_allow_html=True)

# ------------------ RUTAS RELATIVAS ------------------ #
BASE_DIR = os.path.dirname(__file__)
DATA_DIR = os.path.join(BASE_DIR, "..", "data")

RUTA_JSON = os.path.join(DATA_DIR, "ITEMIZADO_jerarquia.json")
RUTA_EXCEL = os.path.join(DATA_DIR, "PRUEBA.xlsx")
HOJA_EXCEL = "protocol"

# ------------------ FUNCIONES ------------------ #
def cargar_jerarquia(ruta):
    if not os.path.exists(ruta):
        st.error(f"No se encontró el archivo JSON en: {ruta}")
        st.stop()
    with open(ruta, "r", encoding="utf-8") as f:
        return json.load(f)

def selector_niveles(diccionario, nivel=1, path=[], etiquetas=[]):
    if not diccionario:
        return path, etiquetas

    opciones = []
    mapa = {}

    for codigo, info in diccionario.items():
        etiqueta = f"{codigo} {info['descripcion']}"
        opciones.append(etiqueta)
        mapa[etiqueta] = (codigo, info)

    opcion_sin_item = "-- No Item --"
    opciones.append(opcion_sin_item)
    mapa[opcion_sin_item] = (f"Sin_Item_N{nivel}", "No Item")

    seleccion = st.selectbox(f"Nivel {nivel}", ["-- Seleccionar --"] + opciones, key=f"nivel_{nivel}")

    if seleccion == opcion_sin_item:
        path.append((f"Sin_Item_N{nivel}", "No Item"))
        etiquetas.append(opcion_sin_item)
        return path, etiquetas

    elif seleccion in mapa:
        codigo, info = mapa[seleccion]
        path.append((codigo, info["descripcion"]))
        etiquetas.append(seleccion)

        if not info.get("subitems"):
            return path, etiquetas

        return selector_niveles(info["subitems"], nivel + 1, path, etiquetas)

    return path, etiquetas

def guardar_fila_excel(path, hoja, fila_dict):
    try:
        wb = load_workbook(path)
        if hoja not in wb.sheetnames:
            st.error(f"La hoja '{hoja}' no existe en el archivo.")
            return
        ws = wb[hoja]
    except Exception as e:
        st.error(f"No se pudo abrir el archivo Excel:\n{e}")
        return

    headers = [cell.value for cell in ws[1]]
    col_map = {col: idx + 1 for idx, col in enumerate(headers)}

    next_row = ws.max_row + 1
    while ws[f"A{next_row}"].value not in [None, ""]:
        next_row += 1

    for campo, valor in fila_dict.items():
        if campo in col_map:
            ws.cell(row=next_row, column=col_map[campo], value=valor)

    try:
        wb.save(path)
        st.success(f"✅ Registro guardado en fila {next_row} de la hoja '{hoja}'.")
    except Exception as e:
        st.error(f"No se pudo guardar el archivo Excel:\n{e}")

def to_int(val): return int(val.strip()) if val.strip() else None
def to_float(val): return float(val.strip()) if val.strip() else None
def limpiar(text): return text.strip().lower() if text.strip() else ""
def formato_protocolo(val): return f"{int(val):05d}" if val.strip().isdigit() else None

# ------------------ SIDEBAR ------------------ #
st.sidebar.image("https://img.icons8.com/fluency/48/000000/construction.png", width=40)
st.sidebar.title("GeoProtocolos")
opcion = st.sidebar.radio("Aplicaciones disponibles", ["📑 Registro de Protocolos"])

# ------------------ APP: Registro de Protocolos ------------------ #
if opcion == "📑 Registro de Protocolos":
    st.title("Registro de Protocolos")
    jerarquia = cargar_jerarquia(RUTA_JSON)
    col_item, col_form = st.columns([1, 2], gap="large")

    with col_item:
        st.markdown("#### Selección Jerárquica")
        codigos_descripciones, etiquetas = selector_niveles(jerarquia)

    with col_form:
        st.markdown("#### Datos del Protocolo")
        with st.form("formulario_manual", clear_on_submit=True):
            col1, col2 = st.columns(2)

            with col1:
                rp_number = st.text_input("Número Recepción Final", "")
                protocol_number = st.text_input("Número Protocolo", "")
                status = st.selectbox("Estado", ["", "Aprobado", "Pendiente", "Rechazado"])
                pk_start = st.text_input("PK Inicio", "")
                pk_end = st.text_input("PK Fin", "")

            with col2:
                work_side = st.selectbox("Lado de Trabajo", ["", "Izquierdo", "Derecho", "Completa"])
                layer_text = st.text_input("Número de Capa", "")
                thickness_text = st.text_input("Espesor (m)", "")
                tag = st.text_input("TAG", "")

            st.markdown("---")
            col_fecha, col_obs = st.columns([1, 2])
            with col_fecha:
                submission_date_rp = st.date_input("Fecha de Envío")
                approval_date_rp = st.date_input("Fecha de Aprobación")
            with col_obs:
                observation_notes = st.text_area("Observaciones", height=115)

            st.markdown("<br>", unsafe_allow_html=True)
            enviar = st.form_submit_button("Guardar registro")

    if enviar:
        fila = {}
        for i in range(5):
            if i < len(codigos_descripciones):
                codigo, descripcion = codigos_descripciones[i]
                fila[f"level_{i+1}"] = f"{codigo} {descripcion}".lower()
            else:
                fila[f"level_{i+1}"] = None

        fila.update({
            "rp_number": formato_protocolo(rp_number),
            "protocol_number": formato_protocolo(protocol_number),
            "status": status.lower() if status else None,
            "pk_start": to_float(pk_start),
            "pk_end": to_float(pk_end),
            "layer": to_int(layer_text),
            "work_side": work_side.lower() if work_side else None,
            "thickness_m": to_float(thickness_text),
            "tag": limpiar(tag),
            "submission_date_rp": submission_date_rp.strftime("%Y-%m-%d"),
            "approval_date_rp": approval_date_rp.strftime("%Y-%m-%d"),
            "observation_notes": observation_notes.strip()
        })

        guardar_fila_excel(RUTA_EXCEL, HOJA_EXCEL, fila)
