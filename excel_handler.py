from openpyxl import load_workbook

def guardar_fila_excel(path, hoja, fila_dict):
    try:
        wb = load_workbook(path)
        if hoja not in wb.sheetnames:
            return False, f"La hoja '{hoja}' no existe."
        ws = wb[hoja]
    except Exception as e:
        return False, f"No se pudo abrir el archivo: {e}"

    headers = [cell.value for cell in ws[1]]
    col_map = {col: idx + 1 for idx, col in enumerate(headers)}

    next_row = ws.max_row + 1
    while ws[f"A{next_row}"].value not in [None, ""]:
        next_row += 1

    for campo, valor in fila_dict.items():
        if campo in col_map:
            ws.cell(row=next_row, column=col_map[campo], value=valor)

    try:
        wb.save(path)
        return True, f"✅ Registro guardado en fila {next_row}"
    except Exception as e:
        return False, f"No se pudo guardar: {e}"
