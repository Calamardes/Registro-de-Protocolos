from fastapi import FastAPI, Request, Form, UploadFile, File
from fastapi.responses import HTMLResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import load_workbook
from datetime import datetime
import shutil
import os
import json

app = FastAPI()

templates = Jinja2Templates(directory="templates")
app.mount("/static", StaticFiles(directory="static"), name="static")

@app.get("/", response_class=HTMLResponse)
def formulario(request: Request):
    return templates.TemplateResponse("formulario.html", {"request": request, "mensaje": ""})

@app.get("/jerarquia")
def obtener_jerarquia():
    json_path = os.path.join("data", "itemizado_acciona.json")
    if not os.path.exists(json_path):
        return JSONResponse(status_code=404, content={"error": "Archivo JSON no encontrado"})
    with open(json_path, "r", encoding="utf-8") as f:
        return json.load(f)

@app.post("/registro", response_class=HTMLResponse)
async def registrar(
    request: Request,
    archivo_excel: UploadFile = File(...),
    rp_number: str = Form(""),
    protocol_number: str = Form(...),
    status: str = Form(""),
    pk_start: str = Form(""),
    pk_end: str = Form(""),
    layer: str = Form(""),
    work_side: str = Form(""),
    thickness_m: str = Form(""),
    tag: str = Form(""),
    submission_date_rp: str = Form(""),
    approval_date_rp: str = Form(""),
    observation_notes: str = Form(""),
    nivel1: str = Form(""),
    nivel2: str = Form(""),
    nivel3: str = Form(""),
    nivel4: str = Form(""),
    nivel5: str = Form("")
):
    if not protocol_number.strip():
        return templates.TemplateResponse("formulario.html", {
            "request": request,
            "mensaje": "❌ Debes ingresar al menos el número de protocolo para guardar."
        })

    # Guardar archivo subido en carpeta temporal
    nombre_excel = archivo_excel.filename
    ruta_guardado = os.path.join("archivos_temporales", nombre_excel)
    os.makedirs("archivos_temporales", exist_ok=True)
    with open(ruta_guardado, "wb") as buffer:
        shutil.copyfileobj(archivo_excel.file, buffer)

    # Abrir Excel y registrar datos
    try:
        wb = load_workbook(ruta_guardado)
        hoja_nombre = "protocol"
        if hoja_nombre not in wb.sheetnames:
            return templates.TemplateResponse("formulario.html", {
                "request": request,
                "mensaje": f"❌ La hoja '{hoja_nombre}' no existe en el archivo."
            })

        ws = wb[hoja_nombre]
        headers = [cell.value for cell in ws[1]]
        col_map = {col: idx + 1 for idx, col in enumerate(headers)}

        nueva_fila = ws.max_row + 1
        while ws[f"A{nueva_fila}"].value not in [None, ""]:
            nueva_fila += 1

        fila = {
            "rp_number": rp_number.strip() or None,
            "protocol_number": protocol_number.strip(),
            "status": status.strip().lower() or None,
            "pk_start": float(pk_start) if pk_start else None,
            "pk_end": float(pk_end) if pk_end else None,
            "layer": int(layer) if layer else None,
            "work_side": work_side.strip().lower() or None,
            "thickness_m": float(thickness_m) if thickness_m else None,
            "tag": tag.strip() or None,
            "submission_date_rp": submission_date_rp or None,
            "approval_date_rp": approval_date_rp or None,
            "observation_notes": observation_notes.strip() or None,
            "level_1": nivel1,
            "level_2": nivel2,
            "level_3": nivel3,
            "level_4": nivel4,
            "level_5": nivel5,
        }

        for campo, valor in fila.items():
            if campo in col_map:
                ws.cell(row=nueva_fila, column=col_map[campo], value=valor)

        wb.save(ruta_guardado)

        return templates.TemplateResponse("formulario.html", {
            "request": request,
            "mensaje": f"✅ Registro guardado exitosamente en la fila {nueva_fila} del archivo '{nombre_excel}'."
        })

    except Exception as e:
        return templates.TemplateResponse("formulario.html", {
            "request": request,
            "mensaje": f"❌ Error al guardar: {str(e)}"
        })
